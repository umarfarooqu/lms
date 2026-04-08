import csv
from datetime import datetime
from io import BytesIO

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import F, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from .forms import BarcodeExcelUploadForm, BarcodeGenerateForm, BookForm, IssueForm
from .models import BarcodeRecord, Book, BookIssue

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None


def librarian_login(request):
    if request.user.is_authenticated:
        return redirect("book_list")

    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        user = authenticate(request, username=username, password=password)

        if user is not None and user.is_staff:
            login(request, user)
            return redirect("book_list")

        messages.error(request, "Invalid librarian credentials.")

    return render(request, "libraryapp/librarian_login.html")


def librarian_logout(request):
    logout(request)
    messages.success(request, "Logged out successfully.")
    return redirect("librarian_login")


@login_required(login_url="librarian_login")
def book_list(request):
    query = request.GET.get("q", "").strip()
    books = Book.objects.all()
    if query:
        books = books.filter(
            Q(title__icontains=query)
            | Q(author__icontains=query)
            | Q(book_code__icontains=query)
        )

    context = {
        "books": books,
        "query": query,
        "book_form": BookForm(),
        "issue_form": IssueForm(),
        "total_books": Book.objects.count(),
        "available_count": Book.objects.aggregate(total=Sum("available_copies"))["total"] or 0,
        "active_issues_count": BookIssue.objects.filter(returned_at__isnull=True).count(),
    }
    return render(request, "libraryapp/book_list.html", context)


@login_required(login_url="librarian_login")
def add_book(request):
    if request.method == "GET":
        return render(request, "libraryapp/add_book.html", {"book_form": BookForm()})

    form = BookForm(request.POST)
    if form.is_valid():
        book = form.save(commit=False)
        book.available_copies = book.total_copies
        book.save()
        rec, _ = BarcodeRecord.objects.get_or_create(
            code=book.book_code,
            defaults={"is_used": True, "book": book},
        )
        if not rec.is_used or rec.book_id != book.id:
            rec.book = book
            rec.is_used = True
            rec.save(update_fields=["book", "is_used"])
        messages.success(request, f"Book '{book.title}' added.")
    else:
        messages.error(request, "Invalid book data.")

    return redirect("add_book")


@login_required(login_url="librarian_login")
def issue_book(request):
    if request.method == "GET":
        return render(request, "libraryapp/issue_book.html", {"issue_form": IssueForm()})

    form = IssueForm(request.POST)
    if not form.is_valid():
        messages.error(request, "Invalid issue request.")
        return redirect("issue_book")

    book = form.cleaned_data["book"]
    if book.available_copies <= 0:
        messages.error(request, "No copies available for this book.")
        return redirect("issue_book")

    student_name = form.cleaned_data.get("student_name", "").strip()
    student_id = form.cleaned_data.get("student_id", "").strip()
    employee_name = form.cleaned_data.get("employee_name", "").strip()
    employee_id = form.cleaned_data.get("employee_id", "").strip()

    if not student_name and not employee_name:
        messages.error(request, "Please fill in either Student or Employee details.")
        return redirect("issue_book")

    BookIssue.objects.create(
        book=book,
        student_name=student_name,
        student_id=student_id,
        employee_name=employee_name,
        employee_id=employee_id,
    )
    Book.objects.filter(pk=book.pk).update(available_copies=F("available_copies") - 1)
    borrower = student_name or employee_name
    messages.success(request, f"Book issued to {borrower}.")
    return redirect("issue_book")


@login_required(login_url="librarian_login")
def return_book(request, issue_id):
    issue = get_object_or_404(BookIssue, pk=issue_id)
    if issue.is_returned:
        messages.info(request, "Book already returned.")
        return redirect("book_list")

    issue.mark_returned()
    Book.objects.filter(pk=issue.book_id).update(available_copies=F("available_copies") + 1)
    messages.success(request, "Book returned successfully.")
    return redirect("book_list")


@login_required(login_url="librarian_login")
def return_book_by_code(request):
    if request.method == "GET":
        return render(request, "libraryapp/return_book.html")

    book_code = request.POST.get("book_code", "").strip().upper()
    student_id = request.POST.get("student_id", "").strip()
    employee_id = request.POST.get("employee_id", "").strip()

    if not book_code or (not student_id and not employee_id):
        messages.error(request, "Book code and either Student ID or Employee ID are required.")
        return redirect("return_book_by_code")

    qs = BookIssue.objects.select_related("book").filter(
        book__book_code=book_code, returned_at__isnull=True
    )
    if student_id:
        issue = qs.filter(student_id=student_id).first()
    else:
        issue = qs.filter(employee_id=employee_id).first()

    if not issue:
        messages.error(request, "Active issue not found for this book code and ID.")
        return redirect("return_book_by_code")

    issue.mark_returned()
    Book.objects.filter(pk=issue.book_id).update(available_copies=F("available_copies") + 1)
    messages.success(request, f"Book {book_code} returned successfully.")
    return redirect("return_book_by_code")


@login_required(login_url="librarian_login")
def student_issues(request):
    student_id = request.GET.get("student_id", "").strip()
    employee_id = request.GET.get("employee_id", "").strip()
    issues = BookIssue.objects.filter(returned_at__isnull=True)
    if student_id:
        issues = issues.filter(student_id=student_id)
    if employee_id:
        issues = issues.filter(employee_id=employee_id)

    context = {
        "student_id": student_id,
        "employee_id": employee_id,
        "issues": issues.select_related("book"),
        "active_issues_count": issues.count(),
    }
    return render(request, "libraryapp/student_issues.html", context)


@login_required(login_url="librarian_login")
def barcode_module(request):
    context = {
        "generate_form": BarcodeGenerateForm(),
        "upload_form": BarcodeExcelUploadForm(),
        "unused_count": BarcodeRecord.objects.filter(is_used=False).count(),
        "recent_codes": BarcodeRecord.objects.order_by("-created_at")[:30],
    }
    return render(request, "libraryapp/barcode_module.html", context)


@login_required(login_url="librarian_login")
def generate_barcodes(request):
    if request.method != "POST":
        return redirect("barcode_module")

    form = BarcodeGenerateForm(request.POST)
    if not form.is_valid():
        messages.error(request, "Invalid quantity.")
        return redirect("barcode_module")

    quantity = form.cleaned_data["quantity"]
    start = BarcodeRecord.get_next_code_number()
    records = [BarcodeRecord(code=str(start + i), is_used=False) for i in range(quantity)]
    BarcodeRecord.objects.bulk_create(records)

    messages.success(request, f"{quantity} barcode(s) generated from {start}.")
    return redirect("barcode_module")


@login_required(login_url="librarian_login")
def export_barcode_sheet(request):
    if Workbook is None:
        messages.error(request, "openpyxl not installed. Run: pip install openpyxl")
        return redirect("barcode_module")

    wb = Workbook()
    ws = wb.active
    ws.title = "Barcode Sheet"
    ws.append(["barcode", "title", "author", "total_copies"])

    for rec in BarcodeRecord.objects.filter(is_used=False).order_by("code"):
        ws.append([rec.code, "", "", ""])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"barcode_sheet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required(login_url="librarian_login")
def upload_barcode_sheet(request):
    if request.method != "POST":
        return redirect("barcode_module")
    if load_workbook is None:
        messages.error(request, "openpyxl not installed. Run: pip install openpyxl")
        return redirect("barcode_module")

    form = BarcodeExcelUploadForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, "Please upload a valid .xlsx file.")
        return redirect("barcode_module")

    workbook = load_workbook(form.cleaned_data["file"], data_only=True)
    sheet = workbook.active

    processed = 0
    created_books = 0
    skipped = 0

    with transaction.atomic():
        for row in sheet.iter_rows(min_row=2, values_only=True):
            barcode = str(row[0]).strip() if row and row[0] is not None else ""
            title = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            author = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            copies_raw = row[3] if len(row) > 3 else 1

            if not barcode:
                continue

            processed += 1
            try:
                total_copies = int(copies_raw) if copies_raw is not None else 1
            except (TypeError, ValueError):
                total_copies = 1
            if total_copies < 1:
                total_copies = 1

            rec = BarcodeRecord.objects.filter(code=barcode).first()
            if not rec or rec.is_used or not title or not author:
                skipped += 1
                continue

            existing = Book.objects.filter(book_code=barcode).first()
            if existing:
                rec.mark_used(existing)
                skipped += 1
                continue

            book = Book.objects.create(
                title=title,
                author=author,
                program_code="18",
                book_code=barcode,
                total_copies=total_copies,
                available_copies=total_copies,
            )
            rec.mark_used(book)
            created_books += 1

    messages.success(
        request,
        f"Sheet processed: {processed}, books created: {created_books}, skipped: {skipped}",
    )
    return redirect("barcode_module")


@login_required(login_url="librarian_login")
def print_barcodes_a4(request):
    records = BarcodeRecord.objects.filter(is_used=False).order_by("code")
    return render(
        request,
        "libraryapp/barcode_print_a4.html",
        {"records": records, "total": records.count()},
    )


@login_required(login_url="librarian_login")
def search_module(request):
    query = request.GET.get("q", "").strip()
    scope = request.GET.get("scope", "all")

    books = Book.objects.none()
    issues = BookIssue.objects.none()
    barcodes = BarcodeRecord.objects.none()

    if query:
        if scope in ("all", "books"):
            books = Book.objects.filter(
                Q(title__icontains=query) | Q(author__icontains=query) | Q(book_code__icontains=query)
            )
        if scope in ("all", "issues"):
            issues = BookIssue.objects.select_related("book").filter(
                Q(student_name__icontains=query)
                | Q(student_id__icontains=query)
                | Q(book__title__icontains=query)
                | Q(book__book_code__icontains=query)
            )
        if scope in ("all", "barcodes"):
            barcodes = BarcodeRecord.objects.select_related("book").filter(Q(code__icontains=query))

    context = {
        "query": query,
        "scope": scope,
        "books": books[:100],
        "issues": issues[:100],
        "barcodes": barcodes[:100],
    }
    return render(request, "libraryapp/search_module.html", context)


@login_required(login_url="librarian_login")
def reports_module(request):
    context = {
        "books_total": Book.objects.count(),
        "books_available": Book.objects.aggregate(total=Sum("available_copies"))["total"] or 0,
        "issues_active": BookIssue.objects.filter(returned_at__isnull=True).count(),
        "issues_returned": BookIssue.objects.filter(returned_at__isnull=False).count(),
        "barcodes_total": BarcodeRecord.objects.count(),
        "barcodes_used": BarcodeRecord.objects.filter(is_used=True).count(),
        "barcodes_unused": BarcodeRecord.objects.filter(is_used=False).count(),
        "latest_books": Book.objects.order_by("-id")[:20],
        "latest_active_issues": BookIssue.objects.select_related("book").filter(returned_at__isnull=True)[:20],
        "latest_returned_issues": BookIssue.objects.select_related("book").filter(returned_at__isnull=False)[:20],
        "latest_barcodes": BarcodeRecord.objects.select_related("book").order_by("-created_at")[:20],
    }
    return render(request, "libraryapp/reports_module.html", context)


@login_required(login_url="librarian_login")
def export_report_csv(request, report_type):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    writer = csv.writer(response)

    if report_type == "books":
        writer.writerow(["book_code", "title", "author", "available_copies", "total_copies"])
        for b in Book.objects.order_by("id"):
            writer.writerow([b.book_code, b.title, b.author, b.available_copies, b.total_copies])
    elif report_type == "active_issues":
        writer.writerow(["issue_id", "book_code", "book_title", "student_name", "student_id", "employee_name", "employee_id", "issued_at"])
        for i in BookIssue.objects.select_related("book").filter(returned_at__isnull=True).order_by("-issued_at"):
            writer.writerow([i.id, i.book.book_code, i.book.title, i.student_name, i.student_id, i.employee_name, i.employee_id, i.issued_at])
    elif report_type == "returned_issues":
        writer.writerow(["issue_id", "book_code", "book_title", "student_name", "student_id", "employee_name", "employee_id", "issued_at", "returned_at"])
        for i in BookIssue.objects.select_related("book").filter(returned_at__isnull=False).order_by("-returned_at"):
            writer.writerow([i.id, i.book.book_code, i.book.title, i.student_name, i.student_id, i.employee_name, i.employee_id, i.issued_at, i.returned_at])
    elif report_type == "barcodes":
        writer.writerow(["code", "status", "book_code", "book_title", "created_at"])
        for r in BarcodeRecord.objects.select_related("book").order_by("-created_at"):
            writer.writerow([r.code, "used" if r.is_used else "unused", r.book.book_code if r.book else "", r.book.title if r.book else "", r.created_at])
    else:
        writer.writerow(["error"])
        writer.writerow(["Invalid report type"])

    return response
